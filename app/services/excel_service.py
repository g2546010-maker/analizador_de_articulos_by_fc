"""
Servicio para generación de archivos Excel con formato institucional.
Exporta artículos académicos con todas sus relaciones y metadatos.
"""
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


class ExcelService:
    """Servicio para generar archivos Excel con formato institucional."""
    
    # Definición de columnas para el reporte
    COLUMNS = [
        ('A', 'ID', 8),
        ('B', 'Título', 50),
        ('C', 'Autores', 40),
        ('D', 'Año', 10),
        ('E', 'Revista', 30),
        ('F', 'Congreso', 30),
        ('G', 'ISSN', 15),
        ('H', 'DOI', 25),
        ('I', 'Tipo de Producción', 20),
        ('J', 'Estado', 15),
        ('K', 'LGAC', 30),
        ('L', 'Propósito', 30),
        ('M', 'Indexaciones', 30),
        ('N', 'País', 15),
        ('O', 'URL', 40),
        ('P', 'Para Currículum', 12),
        ('Q', 'Completo', 10),
        ('R', 'Descripción/Resumen', 60),
    ]
    
    # Colores institucionales
    COLOR_HEADER = 'FF1F4E78'  # Azul institucional
    COLOR_ALT_ROW = 'FFE7E6E6'  # Gris claro para filas alternas
    
    def __init__(self):
        """Inicializa el servicio."""
        self.logger = logger
    
    def generate(self, articulos: List, filename: Optional[str] = None) -> BytesIO:
        """
        Genera archivo Excel con los artículos proporcionados.
        
        Args:
            articulos: Lista de objetos Articulo
            filename: Nombre base del archivo (opcional)
            
        Returns:
            BytesIO con el contenido del archivo Excel
        """
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Artículos Académicos"
            
            # Configurar encabezados
            self._setup_headers(ws)
            
            # Agregar datos
            self._add_data(ws, articulos)
            
            # Aplicar formato
            self._apply_formatting(ws, len(articulos))
            
            # Agregar metadatos
            self._add_metadata(wb, len(articulos))
            
            # Guardar en BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            self.logger.info(f"Excel generado exitosamente con {len(articulos)} artículos")
            return output
            
        except Exception as e:
            self.logger.error(f"Error generando Excel: {str(e)}")
            raise
    
    def _setup_headers(self, ws):
        """Configura los encabezados de las columnas."""
        # Estilos de encabezado
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color=self.COLOR_HEADER, 
                                   end_color=self.COLOR_HEADER, 
                                   fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Aplicar encabezados
        for col_letter, col_name, col_width in self.COLUMNS:
            cell = ws[f'{col_letter}1']
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
            # Ajustar ancho de columna
            ws.column_dimensions[col_letter].width = col_width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
    
    def _add_data(self, ws, articulos: List):
        """Agrega los datos de los artículos al worksheet."""
        for idx, articulo in enumerate(articulos, start=2):
            try:
                # Obtener datos relacionados
                autores = self._get_autores(articulo)
                lgacs = self._get_lgacs(articulo)
                propositos = self._get_propositos(articulo)
                indexaciones = self._get_indexaciones(articulo)
                
                # Llenar fila
                ws[f'A{idx}'] = articulo.id
                ws[f'B{idx}'] = articulo.titulo or ''
                ws[f'C{idx}'] = autores
                ws[f'D{idx}'] = articulo.anio_publicacion or ''
                ws[f'E{idx}'] = articulo.titulo_revista or ''
                ws[f'F{idx}'] = articulo.nombre_congreso or ''
                ws[f'G{idx}'] = articulo.issn or ''
                ws[f'H{idx}'] = articulo.doi or ''
                ws[f'I{idx}'] = articulo.tipo.nombre if articulo.tipo else ''
                ws[f'J{idx}'] = articulo.estado.nombre if articulo.estado else ''
                ws[f'K{idx}'] = lgacs
                ws[f'L{idx}'] = propositos
                ws[f'M{idx}'] = indexaciones
                ws[f'N{idx}'] = articulo.revista.pais.nombre if articulo.revista and articulo.revista.pais else ''
                ws[f'O{idx}'] = articulo.url or ''
                ws[f'P{idx}'] = 'Sí' if articulo.para_curriculum else 'No'
                ws[f'Q{idx}'] = 'Sí' if articulo.completo else 'No'
                ws[f'R{idx}'] = articulo.descripcion or ''
                
            except Exception as e:
                self.logger.warning(f"Error procesando artículo {articulo.id}: {str(e)}")
                continue
    
    def _apply_formatting(self, ws, num_rows: int):
        """Aplica formato a todas las celdas."""
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alineación para datos
        data_alignment = Alignment(vertical='top', wrap_text=True)
        
        # Aplicar a todas las celdas de datos
        for row_idx in range(2, num_rows + 2):
            # Filas alternas
            if row_idx % 2 == 0:
                fill = PatternFill(start_color=self.COLOR_ALT_ROW, 
                                   end_color=self.COLOR_ALT_ROW, 
                                   fill_type='solid')
            else:
                fill = PatternFill()
            
            for col_letter, _, _ in self.COLUMNS:
                cell = ws[f'{col_letter}{row_idx}']
                cell.border = thin_border
                cell.alignment = data_alignment
                cell.fill = fill
                
                # Centrar columnas específicas
                if col_letter in ['A', 'D', 'O', 'P']:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _add_metadata(self, wb, num_articulos: int):
        """Agrega una hoja con metadatos del reporte."""
        ws_meta = wb.create_sheet("Información del Reporte")
        
        # Información del reporte
        metadata = [
            ('Fecha de Generación:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Total de Artículos:', num_articulos),
            ('Sistema:', 'Sistema de Gestión de Artículos Académicos'),
            ('Versión:', '1.0'),
        ]
        
        for idx, (label, value) in enumerate(metadata, start=1):
            ws_meta[f'A{idx}'] = label
            ws_meta[f'B{idx}'] = value
            ws_meta[f'A{idx}'].font = Font(bold=True)
        
        # Ajustar anchos
        ws_meta.column_dimensions['A'].width = 25
        ws_meta.column_dimensions['B'].width = 40
    
    def _get_autores(self, articulo) -> str:
        """Obtiene los autores como string separado por comas."""
        try:
            if not articulo.articulo_autores:
                return ''
            
            autores = []
            for aa in sorted(articulo.articulo_autores, key=lambda x: x.orden):
                autor = aa.autor
                nombre_completo = f"{autor.apellidos}, {autor.nombres}"
                autores.append(nombre_completo)
            
            return '; '.join(autores)
        except Exception as e:
            self.logger.warning(f"Error obteniendo autores: {str(e)}")
            return ''
    
    def _get_lgacs(self, articulo) -> str:
        """Obtiene la LGAC como string."""
        try:
            if articulo.lgac and articulo.lgac.activo:
                return articulo.lgac.nombre
            return ''
        except Exception as e:
            self.logger.warning(f"Error obteniendo LGAC: {str(e)}")
            return ''
    
    def _get_propositos(self, articulo) -> str:
        """Obtiene el propósito como string."""
        try:
            if articulo.proposito and articulo.proposito.activo:
                return articulo.proposito.nombre
            return ''
        except Exception as e:
            self.logger.warning(f"Error obteniendo propósito: {str(e)}")
            return ''
    
    def _get_indexaciones(self, articulo) -> str:
        """Obtiene las indexaciones como string separado por comas."""
        try:
            indexaciones = set()
            
            # Indexaciones de la revista
            if articulo.revista and hasattr(articulo.revista, 'revista_indexaciones'):
                for ri in articulo.revista.revista_indexaciones:
                    if ri.activo and ri.indexacion.activo:
                        indexaciones.add(ri.indexacion.nombre)
            
            # Indexaciones adicionales del artículo
            if hasattr(articulo, 'articulo_indexaciones'):
                for ai in articulo.articulo_indexaciones:
                    if ai.indexacion.activo:
                        indexaciones.add(ai.indexacion.nombre)
            
            return ', '.join(sorted(indexaciones))
        except Exception as e:
            self.logger.warning(f"Error obteniendo indexaciones: {str(e)}")
            return ''
    
    def generate_filename(self, prefix: str = 'articulos') -> str:
        """
        Genera nombre de archivo con timestamp.
        
        Args:
            prefix: Prefijo del nombre del archivo
            
        Returns:
            Nombre de archivo con formato: prefix_YYYYMMDD_HHMMSS.xlsx
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{prefix}_{timestamp}.xlsx"
